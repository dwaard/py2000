from openpyxl import load_workbook

class OpenPyXlReader:

    def __init__(self, filename, edition, start_row=1, title_col=2, artist_col=3, year_col=4, position_col=1):
        self.edition = edition
        self.title_col = title_col
        self.artist_col = artist_col
        self.year_col = year_col
        self.position_col = position_col
        print ("Reading %s..." % filename)
        try:
            workbook = load_workbook(filename, data_only=True)
            self.current_row = start_row
            self.worksheet = workbook.active
        except IOError:
            self.current_row = 0

    def __iter__(self):
        return self

    def __next__(self):
        if self.current_row < 0 or self.current_row > self.worksheet.max_row:
            raise StopIteration
        else:
            while (self.worksheet.cell(self.current_row, 1).value == ''):
                self.current_row += 1
            result = [self.worksheet.cell(self.current_row, self.title_col).value,
                      self.worksheet.cell(self.current_row, self.artist_col).value,
                      int(self.worksheet.cell(self.current_row, self.year_col).value),
                      self.edition,
                       int(self.worksheet.cell(self.current_row, 1).value), ]
            self.current_row += 1
            return result
